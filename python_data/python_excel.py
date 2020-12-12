from openpyxl import Workbook, load_workbook


class pop:
    def pp(self):
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = '身高'
        ws['B1'] = '体重'
        ws['c1'] = '年龄'

        self.hight = [180, 200, 158, 130]
        weight = [180, 150, 100, 90]
        age = [22, 30, 50, 2]
        for i in range(len(self.hight)):
            ws.cell(row=2 + i, column=1, value=self.hight[i])
            ws.cell(row=2 + i, column=2, value=weight[i])
            ws.cell(row=2 + i, column=3, value=age[i])
        wb.save("sample.xlsx")
    def ppo(self):
       la= load_workbook(filename='sample.xlsx') #读取文件
       sheet=la.active #读取页签
       sheet['d1']='备注'
       for i in range(len(self.hight)):
           weight=sheet.cell()